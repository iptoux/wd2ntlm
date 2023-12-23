#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__version__ = "0.2.0"

# Get command line arguments
import argparse

# External module to draw progressbar.
import enlighten

# Default python3 library for logging
import logging

# needed for multithreading
import queue

# We import some files, so we need pathchecks
from pathlib import Path

# Typecheckig style here
from typing import Any

# To work with time and date.
import time
from datetime import datetime, timezone

import threading
import hashlib
import sqlite3

from enum import Enum

# Imports for output formats
import csv
import json
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font



class OFile(Enum):
    JSON = 0
    SQLITE3 = 1
    XLSX = 2
    CSV = 3

# Set some globals
file_name_wordlist: Path = Path()
file_name_ntlmhashes: Path = Path()
file_name_data: dict[str,Any] = {}

data_queue = queue.Queue()
data_processed = set()
data_converted: dict[str,dict[str,str|int|bytes|bytearray|float]] = {}
data_dupes: list = []
data_out_mode: OFile = OFile.SQLITE3

manager = enlighten.get_manager()

thread_local = threading.local()


def sqlConn(db: str) -> None:
    """
    The function `sqlConn` creates a connection to a SQLite database and creates a table named "results"
    if it does not already exist.
    
    :param db: The `db` parameter is a string that represents the name of the database file that you
    want to connect to or create if it doesn't exist
    :type db: str
    """
    conn = sqlite3.Connection(database=db)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item TEXT,
            hash TEXT
        )
    ''')
    conn.commit()
    cursor.close()
    conn.close()

def simpleLog(name: str|None, level: int|str):
    """
    The function `simpleLog` is a Python function that sets up a basic logging configuration and returns
    a logger object.
    
    :param name: The `name` parameter is a string that represents the name of the logger. It is optional
    and can be set to `None` if no specific name is provided
    :type name: str|None
    :param level: The `level` parameter in the `simpleLog` function can accept either an integer or a
    string. It represents the logging level for the logger. The logging levels are defined in the
    `logging` module and include the following options:
    :type level: int|str
    :return: a logger object that can be used for logging messages.
    """

    log_date_format = '%m/%d/%Y %I:%M:%S %p'
    log_formatter = '%(asctime)s - %(name)s.%(funcName)-10s - %(levelname)-8s :: %(message)s'
    log_level = level

    logging.basicConfig(format=log_formatter, datefmt=log_date_format,level=log_level)

    log = logging.getLogger()
    
    if name:
        log.name = name
    else:
        log.name = __name__

    return log

def fileInfo(file: Path):
    """
    The fileInfo function takes a file path as input and returns a dictionary containing information
    about the file, such as its size, number of lines, owner, group, and modification and creation
    timestamps.
    
    :param file: The `file` parameter is of type `Path` and represents the path to the file for which
    you want to retrieve information
    :type file: Path
    :return: The function `fileInfo` returns a dictionary `file_name_data` that contains information
    about the file.
    """
    
    global file_name_data

    # Get file stats and convert tuple to dict
    t = file.stat()

    with open(file) as f:
        num = len(f.readlines())
    f.close()

    if t.st_size > int(100000):    
        st_size:int = int(t.st_size) / 1000000
        b = "Megabytes"
    elif t.st_size > int(10000):    
        st_size:int = int(t.st_size) / 1000
        b = "Kilobytes"
    else:
        st_size:int = int(t.st_size)
        b = "Bytes"

    file_name_data.update({
        "Size": f"{st_size} {b}",
        "SizeRaw": f"{st_size}",
        "Lines": num,
        "Owner": f"(uID){t.st_uid}",
        "Group": f"(gID){t.st_gid}",
        "Modified": datetime.fromtimestamp(int(t.st_mtime), tz=timezone.utc),
    })

    try:
        file_name_data.update({
            "Created": datetime.fromtimestamp(int(t.st_birthtime), tz=timezone.utc),
        })
    except AttributeError:
        file_name_data.update({
            "Created": datetime.fromtimestamp(int(t.st_ctime), tz=timezone.utc),
        })

    
    return file_name_data

def fileLoad(file: Path, threads: int = 1):
    """
    The function `fileLoad` reads the contents of a file, puts each word into a queue, and then converts
    the words using multiple threads.
    
    :param file: The `file` parameter is the path to the file that you want to load. It should be of
    type `Path`
    :type file: Path
    :param threads: The "threads" parameter specifies the number of threads to use for processing the
    file, defaults to 1
    :type threads: int (optional)
    :return: The function `fileLoad` does not explicitly return anything.
    """

    global data_queue, manager, data_converted

    data_converted.update({file.name: {}})

    if int(file_name_data.get('SizeRaw')) >= int(2044): 
        log.info("Reading file data..., files is >BIG< please wait!")
    else:
        log.info("Reading file data...")

    # Init ProgressbarManager
    pbar_rf = manager.counter(total=0, desc='Reading', unit='lines')

    def getFileContents(file): 
        with open(file, 'r') as file_w:
            for line in file_w:
                word = line.strip()            
                data_queue.put(word)
                # slow down here to get now errors
                time.sleep(0.02)
                pbar_rf.update()
    
        file_w.close()
        
    getFileContents(file=file)
    log.info(f"Done. After stripping '\\n' there are {data_queue.qsize()} words in queue")
    
    pbar = manager.counter(total=data_queue.qsize(), desc=f'Converting [{threads} threads]', unit='words')

    workerMgr(thread_count=threads,word_queue=data_queue,progressbar=pbar)
    
    return

def fileSave(format):

    def saveJson():
        with open(f"{file_name_ntlmhashes}.json",'w') as f:
            f.write(json.dumps(data_converted,indent=4))
        f.close()
    def saveCSV():
        with open(f"{file_name_ntlmhashes}.csv",'w') as f:
            w = csv.writer(f)
            rows = data_converted.get(list(data_converted.keys())[0])
            w.writerows(rows.items())
        f.close()
    def saveXlsx():        
        xlsx = Workbook()
        sheet = xlsx.active
        sheet.title = 'wd2ntlm'
        
        headers = ['HASH','WORD']
        sheet.append(headers)

        row=2
        sheet.cell(row=row-1, column=1).font = Font(bold=True)
        sheet.cell(row=row-1, column=2).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(1)].width = 50
        sheet.column_dimensions[get_column_letter(2)].width = 20

        for row, (hash, word) in enumerate(data_converted.get(list(data_converted.keys())[0]).items(), start=2):
             sheet[f"A{row}"] = hash
             sheet[f"B{row}"] = word

        xlsx.save(f"{file_name_ntlmhashes}.xlsx")

    match format:
        case OFile.JSON:
            saveJson()
            log.debug("Called saveJson()")
        case OFile.XLSX:
            saveXlsx()
            log.debug("Called saveXlsx()")
        case OFile.CSV:
            saveCSV()
            log.debug("Called saveCsv()")

            pass
    return

def worker(queue, thread_id, progressbar, data_processed):
    """
    The `worker` function processes items from a queue, calculates the MD4 hash of each item, saves the
    item and its hash into a database, and updates a progress bar.
    
    :param queue: The `queue` parameter is a queue object that contains the items to be processed by the
    worker function. Each worker thread will retrieve items from this queue and process them
    :param thread_id: The `thread_id` parameter is used to identify the thread that the worker function
    is running on. It can be any value that uniquely identifies the thread, such as an integer or a
    string. This parameter is useful for logging and debugging purposes to track the progress of each
    thread
    :param progressbar: The `progressbar` parameter is a progress bar object that is used to track the
    progress of the worker thread. It is updated each time an item is processed from the queue
    :param data_processed: The `data_processed` parameter is a set that keeps track of the items that
    have already been processed by the worker. It is used to avoid processing duplicate items from the
    queue
    """
    
    global data_converted
    log.info(f"[T|{thread_id}]: I'm alive!")
    workerDb(thread_id)

    while not queue.empty():
        item = queue.get()
        if item not in data_processed:

            log.debug(f"[T|{thread_id}]: Processing -> {item}")
            
            input_bytes = item.encode('utf-16le')  # NTLM uses UTF-16LE encoding

            # Calculate the MD4 hash of the input bytes
            md4_hash = hashlib.new('md4')
            md4_hash.update(input_bytes)
            hash_result = md4_hash.digest()

            log.debug(f"[T|{thread_id}]: Hash \t-> {hash_result.hex()}")

            if data_out_mode != OFile.SQLITE3:
                data_converted[file_name_wordlist.name].update({
                    hash_result.hex():item
                })

            log.debug(f"[T|{thread_id}]: Saving into database.")
            thread_local.db_cursor.execute('INSERT INTO results (item, hash) VALUES (?, ?)', (item, hash_result.hex()))
            thread_local.db_connection.commit()
            data_processed.add(item)
            log.debug(f"[T|{thread_id}]: -------------------------------------------")
        else:
            log.debug(f"[T|{thread_id}]: Duplicate -> {item}")
            data_dupes.append(item)
        
        time.sleep(0.02)
        progressbar.update()
        queue.task_done()

def workerDb(thread_id):
    """
    The function `workerDb` opens a new connection to a SQLite database and assigns it to a thread-local
    variable.
    
    :param thread_id: The `thread_id` parameter is an identifier for the current thread. It is used to
    differentiate between different threads when logging messages or performing other operations
    specific to a particular thread
    """
    if hasattr(thread_local, 'db_connection') and thread_local.db_connection:
        log.debug(f"[T|{thread_id}]: closing stalled connection.")
        thread_local.db_connection.close()

    log.debug(f"[T|{thread_id}]: open new connection to database")
    thread_local.db_connection = sqlite3.connect(f"{file_name_ntlmhashes.name}.sqlite3")
    thread_local.db_cursor = thread_local.db_connection.cursor()

def workerMgr(thread_count, word_queue,progressbar):
    """
    The function `workerMgr` starts a specified number of worker threads to process words from a queue,
    and waits for all threads to finish before closing any open database connections.
    
    :param thread_count: The `thread_count` parameter specifies the number of worker threads to create.
    These threads will be responsible for processing the words in the `word_queue`
    :param word_queue: The `word_queue` parameter is a queue that holds the words that need to be
    processed by the worker threads. Each worker thread will take a word from the queue and process it
    :param progressbar: The `progressbar` parameter is a variable that is passed to the `workerMgr`
    function. It is used to track the progress of the workers as they process the words in the
    `word_queue`. It is likely an instance of a progress bar class or module that provides methods for
    updating and displaying the
    """

    global data_processed

    log.info(f"Starting {thread_count} workers for {word_queue.qsize()} words.")

    # Create a list to hold the thread objects
    threads = []

    if thread_count >=15:
        log.warning(f"THREAD_COUNT >= 15/{thread_count} You should not do that! Waiting 10sec...")
        time.sleep(10)

    # Start the threads
    for i in range(thread_count):
        thread = threading.Thread(target=worker, args=(word_queue, i+1,progressbar, data_processed))
        
        thread.start()
        threads.append(thread)

    # Wait for all threads to finish
    for thread in threads:
        thread.join()

    if hasattr(thread_local, 'db_connection') and thread_local.db_connection:
        thread_local.db_connection.close()

    log.info("All workers finished!")

def main(file:str, threads:int = 1, debug = False):
    """
    The main function takes a file name, number of threads, and a debug flag as input, and performs
    various operations on the file, including loading data, processing duplicates, and saving the data
    in different formats.
    
    :param file: The `file` parameter is a string that represents the path to the input file. It is
    required for the function to run
    :type file: str
    :param threads: The `threads` parameter specifies the number of threads to use during the conversion
    process. It determines how many parallel processes will be used to convert the data. By default, it
    is set to 1, meaning no parallel processing will be used, defaults to 1
    :type threads: int (optional)
    :param debug: The `debug` parameter is a boolean flag that determines whether debug logging should
    be enabled or not. If `debug` is set to `True`, then debug logging will be enabled. If `debug` is
    set to `False` (default), then debug logging will be disabled, defaults to False (optional)
    :return: The function `main` does not explicitly return anything.
    """

    global file_name_wordlist 
    file_name_wordlist = Path(file)
    
    if debug:
        log.setLevel(logging.DEBUG)
        log.info(f"Debug: Enabled by argument.")

    log.info(f"File: {file_name_wordlist}")
    log.info(f"OutFile: {file_name_ntlmhashes}.{data_out_mode.name.lower()}")    
    if not Path.exists(file_name_wordlist):
        log.critical("File not found! Exit(1)")
        exit(1)


    log.info("Init SQLite3 for storage.")
    sqlConn(db=f"{file_name_ntlmhashes.name}.sqlite3")

    file_info = fileInfo(file_name_wordlist)
    for key,val in file_info.items():
        if key == "Size" or key == "Lines":
            log.info(f"\t{key} => {val}")
        else:
            log.debug(f"\t{key} => {val}")

    log.info(f"Threads to use: {threads} (during convert)")

    fileLoad(file_name_wordlist, threads)
    log.info(f"Done, [{len(data_processed)}/{file_name_data['Lines']}] Skipped {len(data_dupes)} duplicates")
    log.debug(f"Dupes: {data_dupes}")

    log.info(f"Saving data to file: {file_name_ntlmhashes}.{data_out_mode.name.lower()}")

    if data_out_mode == OFile.JSON:
        fileSave(OFile.JSON)
    elif data_out_mode == OFile.CSV:
        fileSave(OFile.CSV)        
    elif data_out_mode == OFile.XLSX:
        fileSave(OFile.XLSX)
    return


if __name__ == "__main__":

    # Set name
    __name__ = "wd2ntlm"
    
    # Init simple logger
    log = simpleLog(name=__name__,level=logging.INFO)
    
    # Load function to read arguments
    parser = argparse.ArgumentParser(prog=__name__,
                                     description='Convert passwords from a file to ntlm hashes. \
                                        The output can be selected between json, sqlite3(default)\
                                        , xlsx and csv.',)

    general_group = parser.add_argument_group('General Options')
    style_group = parser.add_argument_group('Output Options')
    nerd_group = parser.add_argument_group('Nerd Options')

    nerd_group.add_argument('-d','--debug',
                        action='store_true',
                        help='enable debug output')
    general_group.add_argument('-f','--file',
                        metavar="FILE",
                        type=str,
                        required=True,
                        help='the file to read and convert')
    general_group.add_argument('-o','--out',
                        metavar="FILE",
                        type=str,
                        required=True,
                        help='the output file to write hashes')
    style_group.add_argument('--csv',
                        action='store_true',
                        required=False,
                        help='output file should be in csv fromat.')
    style_group.add_argument('--json',
                        action='store_true',
                        required=False,
                        help='output file should be in JSON fromat.')
    style_group.add_argument('--xlsx',
                        action='store_true',
                        required=False,
                        help='output file should be in Excel fromat.')
    general_group.add_argument('-t','--threads',
                        type=int,
                        required=False,
                        default=1,
                        metavar="N",
                        help='how many threads should be used?')  
    general_group.add_argument('-v','--version',
                        action='version',
                        version=f'%(prog)s {__version__}')

    args = parser.parse_args()

    if args.csv:
        data_out_mode: OFile = OFile.CSV
    elif args.json:
        data_out_mode: OFile = OFile.JSON
    elif args.xlsx:
        data_out_mode: OFile = OFile.XLSX

    file_name_ntlmhashes = Path(f"{args.out}")

    # Init main function with argument.
    main(file=args.file, debug=args.debug, threads=args.threads)